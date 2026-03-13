#!/usr/bin/env python3
"""
ops_helper.py - VIPKID 运营后台商品包操作辅助工具

用法：
  python3 ops_helper.py auth                                                   # 验证 token 是否有效
  python3 ops_helper.py refresh-token [--timeout 300] [--headless]             # 通过 Playwright 独立登录并自动刷新 token
  python3 ops_helper.py refresh-token --mode cdp --port 9222                   # 可选：通过已开启 CDP 的 Chrome 读取 token
  python3 ops_helper.py list [名称关键词]                                       # 列出商品包
  python3 ops_helper.py detail <packageId>                                     # 查看详情
  python3 ops_helper.py coupon-limit <packageId>                               # 查看优惠券限制
  python3 ops_helper.py inventory <packageId>                                  # 查看库存
  python3 ops_helper.py update-stock <packageId> add <n>                       # 追加库存 n
  python3 ops_helper.py update-stock <packageId> subtract <n>                  # 扣减库存 n
  python3 ops_helper.py update-stock <packageId> infinity                      # 设为不限制库存
  python3 ops_helper.py from-excel <file.xlsx>                                 # 解析 Excel 批量数据
  python3 ops_helper.py batch-create <file.json>                               # 批量新建（需二次确认）
"""
import argparse
import json
import os
import subprocess
import sys
import urllib.error
import urllib.request
from pathlib import Path

CONFIG_PATH = Path.home() / ".vipkid-ops" / "config.json"
DEFAULT_BASE_URL = "https://sa-manager.lionabc.com"
DEFAULT_CR_CODE = "sa"
DEFAULT_CDP_PORT = 9222
DEFAULT_REFRESH_MODE = "playwright"
DEFAULT_LOGIN_TIMEOUT_SECONDS = 300
DEFAULT_PLAYWRIGHT_PROFILE_DIR = CONFIG_PATH.parent / "playwright-profile"


def load_config(required=False):
    if not CONFIG_PATH.exists():
        if required:
            print(f"[ERROR] 配置文件不存在：{CONFIG_PATH}")
            print("请先创建配置文件，或直接运行 refresh-token 自动写入 token。")
            print('建议格式：{"base_url": "https://sa-manager.lionabc.com", "token": ""}')
            sys.exit(1)
        return {"base_url": DEFAULT_BASE_URL, "cr_code": DEFAULT_CR_CODE}

    with open(CONFIG_PATH, encoding="utf-8") as file_obj:
        config = json.load(file_obj)

    config.setdefault("base_url", DEFAULT_BASE_URL)
    config.setdefault("cr_code", DEFAULT_CR_CODE)
    return config


def save_config(config):
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_PATH, "w", encoding="utf-8") as file_obj:
        json.dump(config, file_obj, indent=2, ensure_ascii=False)
        file_obj.write("\n")


def require_auth_config():
    config = load_config(required=True)
    token = str(config.get("token", "")).strip()
    if token:
        return config

    print(f"[ERROR] 配置文件缺少 token：{CONFIG_PATH}")
    print("可先运行以下命令自动获取：")
    print("  python3 ~/.claude/skills/vipkid-ops/scripts/ops_helper.py refresh-token")
    print("首次运行会打开一个独立 Chrome 窗口，你登录后脚本会自动读取 intlAuthToken。")
    sys.exit(1)


def build_headers():
    config = require_auth_config()
    token = config["token"].strip()
    base_url = config["base_url"]
    cr_code = config.get("cr_code", DEFAULT_CR_CODE)

    headers = {
        "intlAuthToken": token,
        "Cookie": f"intlAuthToken={token}",
        "vk-cr-code": cr_code,
        "vk-language-code": "zh-cn",
        "app-code": "international-mgt",
        "Referer": f"{base_url}/hobbit/product/productpackagemanage/",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    return config, headers


def api_get(path, params=None):
    config, headers = build_headers()
    url = f"{config['base_url']}/rest{path}"
    if params:
        from urllib.parse import urlencode

        url += "?" + urlencode(params)

    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as error:
        try:
            body = json.loads(error.read().decode())
        except Exception:
            body = {"http_error": error.code, "reason": str(error.reason)}
        return body
    except urllib.error.URLError as error:
        return {"network_error": str(error.reason)}


def api_post(path, data):
    config, headers = build_headers()
    url = f"{config['base_url']}/rest{path}"
    req = urllib.request.Request(url, data=json.dumps(data).encode(), headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as error:
        try:
            body = json.loads(error.read().decode())
        except Exception:
            body = {"http_error": error.code}
        return body
    except urllib.error.URLError as error:
        return {"network_error": str(error.reason)}


def parse_refresh_args(args):
    parser = argparse.ArgumentParser(
        prog="python3 ops_helper.py refresh-token",
        description="Refresh intlAuthToken via Playwright or CDP.",
    )
    parser.add_argument(
        "--mode",
        choices=("playwright", "cdp"),
        default=os.environ.get("VIPKID_TOKEN_REFRESH_MODE", DEFAULT_REFRESH_MODE),
        help="Refresh strategy. Defaults to Playwright.",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=int(os.environ.get("VIPKID_CHROME_CDP_PORT", DEFAULT_CDP_PORT)),
        help="CDP port for --mode cdp.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=int(os.environ.get("VIPKID_TOKEN_WAIT_TIMEOUT", DEFAULT_LOGIN_TIMEOUT_SECONDS)),
        help="Seconds to wait for Playwright login.",
    )
    parser.add_argument(
        "--profile-dir",
        default=str(DEFAULT_PLAYWRIGHT_PROFILE_DIR),
        help="Dedicated Playwright profile directory.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Playwright in headless mode. Mainly useful for automated verification.",
    )
    return parser.parse_args(args)


def run_refresh_command(command, timeout_seconds):
    try:
        return subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=False,
        )
    except FileNotFoundError as error:
        print(f"[ERROR] 未检测到命令：{error.filename}")
        return None
    except subprocess.TimeoutExpired:
        print("❌ 自动获取 Token 超时。")
        return None


def refresh_token(mode=DEFAULT_REFRESH_MODE, port=DEFAULT_CDP_PORT, timeout_seconds=DEFAULT_LOGIN_TIMEOUT_SECONDS, profile_dir=DEFAULT_PLAYWRIGHT_PROFILE_DIR, headless=False):
    config = load_config(required=False)
    base_url = config.get("base_url", DEFAULT_BASE_URL)
    profile_dir = Path(profile_dir).expanduser()

    if mode == "playwright":
        script_path = Path(__file__).with_name("get_token_via_playwright.py")
        if not script_path.exists():
            print(f"[ERROR] 未找到 Playwright 脚本：{script_path}")
            return False

        command = [
            sys.executable,
            str(script_path),
            "--url",
            base_url,
            "--cookie",
            "intlAuthToken",
            "--timeout",
            str(timeout_seconds),
            "--profile-dir",
            str(profile_dir),
        ]
        if headless:
            command.append("--headless")

        print(f"正在打开独立 Chrome 窗口，使用专用登录目录：{profile_dir}")
        print("请在弹出的浏览器里完成登录，脚本会在检测到 intlAuthToken 后自动关闭窗口。")
        result = run_refresh_command(command, timeout_seconds + 90)
        failure_hints = [
            "  1. 是否已在弹出的独立 Chrome 窗口中完成登录",
            "  2. 登录成功后 Cookie 中是否存在 intlAuthToken",
            f"  3. 如需重置专用登录态，可删除目录：{profile_dir}",
        ]
        success_message = f"✅ 已通过 Playwright 读取 intlAuthToken，并写入 {CONFIG_PATH}"
    else:
        script_path = Path(__file__).with_name("get_token_via_cdp.js")
        if not script_path.exists():
            print(f"[ERROR] 未找到 CDP 脚本：{script_path}")
            return False

        command = [
            "node",
            str(script_path),
            "--port",
            str(port),
            "--url",
            base_url,
            "--cookie",
            "intlAuthToken",
        ]
        result = run_refresh_command(command, 30)
        failure_hints = [
            f"  1. Chrome 是否已使用 --remote-debugging-port={port} 启动",
            "  2. 是否已在 Chrome 中登录 sa-manager.lionabc.com",
            "  3. Cookie 中是否存在 intlAuthToken",
        ]
        success_message = f"✅ 已通过 Chrome CDP 读取 intlAuthToken，并写入 {CONFIG_PATH}"

    if result is None:
        return False

    token = result.stdout.strip()
    error_output = result.stderr.strip()
    if result.returncode != 0 or not token:
        print("❌ 自动获取 Token 失败。")
        if error_output:
            print(error_output)
        print("可检查：")
        for hint in failure_hints:
            print(hint)
        return False

    config["base_url"] = base_url
    config["cr_code"] = config.get("cr_code", DEFAULT_CR_CODE)
    config["token"] = token
    save_config(config)
    print(success_message)
    return True


def check_auth():
    resp = api_get("/auth/api/auth/userAuthInfo")
    if resp.get("code") == 200 and resp.get("data"):
        role_list = resp["data"].get("roleList", [])
        print(f"✅ 认证有效，角色：{role_list}")
        return True

    print("❌ Token 已过期或无效！")
    if resp.get("network_error"):
        print(f"   网络错误：{resp['network_error']}")
    elif resp.get("http_error"):
        print(f"   HTTP 错误：{resp['http_error']}")
    elif resp.get("msg"):
        print(f"   服务端返回：{resp['msg']}")
    print("   可先尝试自动刷新：")
    print("   python3 ~/.claude/skills/vipkid-ops/scripts/ops_helper.py refresh-token")
    print("   该命令会打开一个独立 Chrome 窗口，你登录后脚本会自动读取 intlAuthToken。")
    print("   若仍失败，再手动从 Chrome → F12 → Application → Cookies → sa-manager.lionabc.com → intlAuthToken 复制 Value")
    return False


def list_packages(name=None, page=0, limit=10):
    params = {"queryType": "Product_Package", "start": page * limit, "limit": limit}
    if name:
        params["productName"] = name
    resp = api_get("/international/api/product/list/", params)
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return []

    total = resp["data"]["total"]
    items = resp["data"]["data"]
    print(f"共 {total} 个商品包，本页 {len(items)} 个：")
    print(f"{'ID':<8} {'名称':<35} {'状态':<6} {'售价':<15} {'创建人':<12}")
    print("-" * 80)
    for item in items:
        status = "上架" if item.get("status") == 1 else "下架"
        price = item.get("realPrice", "?")
        if isinstance(price, str) and len(price) > 20:
            price = price[:17] + "..."
        print(f"{item['id']:<8} {item['name'][:34]:<35} {status:<6} {str(price):<15} {item.get('createName', '?'):<12}")
    return items


def get_detail(package_id):
    resp = api_get("/international/api/product/detail", {"productId": package_id, "type": "Package"})
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return None
    data = resp["data"]
    print(json.dumps(data, ensure_ascii=False, indent=2))
    return data


def get_coupon_limit(package_id):
    resp = api_get("/international/api/product/getCouponLimit/", {"productId": package_id})
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return None
    data = resp["data"]
    print(f"商品包 {package_id} 优惠券限制：")
    print(f"  数量限制：{data.get('couponLimitNum', '无限制')}")
    print(f"  金额比例限制：{data.get('couponLimitRate', '无限制')}%")
    return data


def get_inventory(package_id):
    resp = api_get("/international/order-service/api/product/package/stock", {"packageId": package_id})
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return None
    data = resp["data"]
    print(f"商品包 {package_id} 库存信息：")
    stock_limit = data.get("stockLimit")
    print(f"  是否限制：{'限制' if stock_limit else '不限制'}")
    print(f"  库存总量：{data.get('stockNum', '-')}")
    print(f"  待支付：{data.get('toBePaidNum', '-')}")
    print(f"  退款中：{data.get('inRefundNum', '-')}")
    print(f"  已售：{data.get('paymentNum', '-')}")
    return data


def update_inventory(package_id, operate_type, operate_num, product_type_id=3):
    """
    Update package inventory.
    operate_type: "add" | "subtract" | None for unlimited stock.
    operate_num: numeric delta when operate_type is not None.
    """
    if operate_type is None:
        payload = {"packageId": package_id, "stockLimit": False}
    elif operate_type in ("add", "subtract"):
        payload = {
            "packageId": package_id,
            "stockLimit": True,
            "operateType": operate_type,
            "operateNum": operate_num,
            "productTypeId": product_type_id,
        }
    else:
        print(f"[ERROR] operateType 只接受 'add' / 'subtract' / None，当前值：{operate_type}")
        return None

    resp = api_post("/international/order-service/api/product/package/updateStock", payload)
    if resp.get("code") == 200:
        print(f"✅ 库存更新成功，当前库存：{resp.get('data', {}).get('stock', '-')}")
    else:
        print(f"❌ 库存更新失败：{resp.get('msg', resp)}")
    return resp


def parse_excel(filepath):
    """Parse Excel file to extract product package data."""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"Excel 列：{headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(headers, row)))

    print(f"共 {len(rows)} 行数据")
    return rows


def batch_create(data_list, dry_run=True):
    """Batch create product packages with confirmation."""
    print(f"\n{'[DRY RUN] ' if dry_run else ''}准备新建 {len(data_list)} 个商品包：")
    print(f"{'ID':<6} {'名称':<40} {'原价':<12} {'售价':<12}")
    print("-" * 70)
    for index, item in enumerate(data_list):
        print(f"{index + 1:<6} {str(item.get('name', '?'))[:39]:<40} {item.get('originPrice', '?'):<12} {item.get('realPrice', '?'):<12}")

    if dry_run:
        print("\n[DRY RUN 模式] 以上是将要新建的数据，请检查后手动调用 batch_create(data, dry_run=False)")
        return

    print("\n确认新建以上商品包？(y/N): ", end="")
    answer = input().strip().lower()
    if answer != "y":
        print("已取消")
        return

    results = {"success": [], "failed": []}
    for index, item in enumerate(data_list):
        payload = {
            **item,
            "content": item.get("content", []),
            "multiCurrencyPricingData": item.get("multiCurrencyPricingData", []),
        }
        resp = api_post("/international/order-service/api/product/new", payload)
        if resp.get("code") == 200:
            new_id = resp.get("data", {}).get("id", "?")
            print(f"  ✅ [{index + 1}] {item.get('name')} → ID={new_id}")
            results["success"].append({"name": item.get("name"), "id": new_id})
        else:
            print(f"  ❌ [{index + 1}] {item.get('name')} → {resp.get('msg', resp)}")
            results["failed"].append({"name": item.get("name"), "error": resp.get("msg")})

    print(f"\n完成：成功 {len(results['success'])}，失败 {len(results['failed'])}")
    if results["failed"]:
        print("失败列表：")
        for failed in results["failed"]:
            print(f"  {failed['name']}: {failed['error']}")
    return results


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    cmd = sys.argv[1]

    if cmd == "auth":
        check_auth()
    elif cmd == "refresh-token":
        refresh_args = parse_refresh_args(sys.argv[2:])
        refresh_token(
            mode=refresh_args.mode,
            port=refresh_args.port,
            timeout_seconds=refresh_args.timeout,
            profile_dir=refresh_args.profile_dir,
            headless=refresh_args.headless,
        )
    elif cmd == "list":
        name = sys.argv[2] if len(sys.argv) > 2 else None
        list_packages(name)
    elif cmd == "detail":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py detail <packageId>")
            sys.exit(1)
        get_detail(int(sys.argv[2]))
    elif cmd == "coupon-limit":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py coupon-limit <packageId>")
            sys.exit(1)
        get_coupon_limit(int(sys.argv[2]))
    elif cmd == "inventory":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py inventory <packageId>")
            sys.exit(1)
        get_inventory(int(sys.argv[2]))
    elif cmd == "update-stock":
        if len(sys.argv) < 4:
            print("用法: python3 ops_helper.py update-stock <packageId> add|subtract|infinity [num]")
            sys.exit(1)
        package_id = int(sys.argv[2])
        operation = sys.argv[3]
        if operation == "infinity":
            update_inventory(package_id, None, None)
        elif operation in ("add", "subtract"):
            if len(sys.argv) < 5:
                print(f"用法: python3 ops_helper.py update-stock <packageId> {operation} <num>")
                sys.exit(1)
            update_inventory(package_id, operation, int(sys.argv[4]))
        else:
            print(f"[ERROR] 操作类型只支持 add / subtract / infinity，收到：{operation}")
    elif cmd == "from-excel":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py from-excel <file.xlsx>")
            sys.exit(1)
        rows = parse_excel(sys.argv[2])
        print(json.dumps(rows[:3], ensure_ascii=False, indent=2, default=str))
    elif cmd == "batch-create":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py batch-create <file.json>")
            sys.exit(1)
        with open(sys.argv[2], encoding="utf-8") as file_obj:
            data = json.load(file_obj)
        batch_create(data, dry_run=True)
    else:
        print(f"未知命令: {cmd}")
        print(__doc__)
