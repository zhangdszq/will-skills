#!/usr/bin/env python3
"""
ops_helper.py - VIPKID 运营后台商品包操作辅助工具

用法：
  python3 ops_helper.py auth                             # 验证 token 是否有效
  python3 ops_helper.py list [名称关键词]                 # 列出商品包
  python3 ops_helper.py detail <packageId>               # 查看详情
  python3 ops_helper.py coupon-limit <packageId>         # 查看优惠券限制
  python3 ops_helper.py inventory <packageId>            # 查看库存
  python3 ops_helper.py update-stock <packageId> add <n>      # 追加库存 n
  python3 ops_helper.py update-stock <packageId> subtract <n> # 扣减库存 n
  python3 ops_helper.py update-stock <packageId> infinity     # 设为不限制库存
  python3 ops_helper.py from-excel <file.xlsx>           # 解析 Excel 批量数据
  python3 ops_helper.py batch-create <file.json>         # 批量新建（需二次确认）
"""
import json
import sys
import os
import urllib.request
import urllib.error
from pathlib import Path

CONFIG_PATH = Path.home() / ".vipkid-ops" / "config.json"

# Load config
def load_config():
    if not CONFIG_PATH.exists():
        print(f"[ERROR] 配置文件不存在：{CONFIG_PATH}")
        print("请创建配置文件，格式：")
        print('  {"base_url": "https://sa-manager.lionabc.com", "token": "<从浏览器Cookie复制>"}')
        sys.exit(1)
    with open(CONFIG_PATH) as f:
        return json.load(f)

config = load_config()
TOKEN = config["token"]
BASE_URL = config["base_url"]
CR_CODE = config.get("cr_code", "sa")

HEADERS = {
    "intlAuthToken": TOKEN,
    "Cookie": f"intlAuthToken={TOKEN}",
    "vk-cr-code": CR_CODE,
    "vk-language-code": "zh-cn",
    "app-code": "international-mgt",
    "Referer": f"{BASE_URL}/hobbit/product/productpackagemanage/",
    "Accept": "application/json",
    "Content-Type": "application/json",
}

def api_get(path, params=None):
    url = f"{BASE_URL}/rest{path}"
    if params:
        from urllib.parse import urlencode
        url += "?" + urlencode(params)
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        try: body = json.loads(e.read().decode())
        except: body = {"http_error": e.code, "reason": str(e.reason)}
        return body

def api_post(path, data):
    url = f"{BASE_URL}/rest{path}"
    req = urllib.request.Request(url, data=json.dumps(data).encode(), headers=HEADERS, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        try: body = json.loads(e.read().decode())
        except: body = {"http_error": e.code}
        return body

def check_auth():
    resp = api_get("/auth/api/auth/userAuthInfo")
    if resp.get("code") == 200 and resp.get("data"):
        role_list = resp["data"].get("roleList", [])
        print(f"✅ 认证有效，角色：{role_list}")
        return True
    print("❌ Token 已过期或无效！")
    print("   请从浏览器重新复制 intlAuthToken Cookie 到 ~/.vipkid-ops/config.json")
    print("   步骤：Chrome → F12 → Application → Cookies → sa-manager.lionabc.com → intlAuthToken → 复制 Value")
    return False

def list_packages(name=None, page=0, limit=10):
    params = {"queryType": "Product_Package", "start": page * limit, "limit": limit}
    if name:
        params["productName"] = name
    resp = api_get("/international/api/product/list/", params)
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return []
    total = resp["data"]["total"]
    items = resp["data"]["data"]
    print(f"共 {total} 个商品包，本页 {len(items)} 个：")
    print(f"{'ID':<8} {'名称':<35} {'状态':<6} {'售价':<15} {'创建人':<12}")
    print("-" * 80)
    for item in items:
        status = "上架" if item.get("status") == 1 else "下架"
        price = item.get("realPrice", "?")
        # realPrice might be multi-currency or single
        if isinstance(price, str) and len(price) > 20:
            price = price[:17] + "..."
        print(f"{item['id']:<8} {item['name'][:34]:<35} {status:<6} {str(price):<15} {item.get('createName','?'):<12}")
    return items

def get_detail(package_id):
    resp = api_get("/international/api/product/detail", {"productId": package_id, "type": "Package"})
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return None
    data = resp["data"]
    print(json.dumps(data, ensure_ascii=False, indent=2))
    return data

def get_coupon_limit(package_id):
    resp = api_get("/international/api/product/getCouponLimit/", {"productId": package_id})
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return None
    data = resp["data"]
    print(f"商品包 {package_id} 优惠券限制：")
    print(f"  数量限制：{data.get('couponLimitNum', '无限制')}")
    print(f"  金额比例限制：{data.get('couponLimitRate', '无限制')}%")
    return data

def get_inventory(package_id):
    resp = api_get("/international/order-service/api/product/package/stock", {"packageId": package_id})
    if resp.get("code") != 200:
        print(f"[ERROR] {resp}")
        return None
    data = resp["data"]
    print(f"商品包 {package_id} 库存信息：")
    stock_limit = data.get("stockLimit")
    print(f"  是否限制：{'限制' if stock_limit else '不限制'}")
    print(f"  库存总量：{data.get('stockNum', '-')}")
    print(f"  待支付：{data.get('toBePaidNum', '-')}")
    print(f"  退款中：{data.get('inRefundNum', '-')}")
    print(f"  已售：{data.get('paymentNum', '-')}")
    return data


def update_inventory(package_id, operate_type, operate_num, product_type_id=3):
    """
    更新库存。
    operate_type: "add" 追加 | "subtract" 扣减 | None 设置为不限制
    operate_num: 操作数量（仅 operate_type 非 None 时有效）
    """
    if operate_type is None:
        payload = {"packageId": package_id, "stockLimit": False}
    elif operate_type in ("add", "subtract"):
        payload = {
            "packageId": package_id,
            "stockLimit": True,
            "operateType": operate_type,
            "operateNum": operate_num,
            "productTypeId": product_type_id,
        }
    else:
        print(f"[ERROR] operateType 只接受 'add' / 'subtract' / None，当前值：{operate_type}")
        return None
    resp = api_post("/international/order-service/api/product/package/updateStock", payload)
    if resp.get("code") == 200:
        print(f"✅ 库存更新成功，当前库存：{resp.get('data', {}).get('stock', '-')}")
    else:
        print(f"❌ 库存更新失败：{resp.get('msg', resp)}")
    return resp

def parse_excel(filepath):
    """Parse Excel file to extract product package data"""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"Excel 列：{headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # skip empty rows
            rows.append(dict(zip(headers, row)))

    print(f"共 {len(rows)} 行数据")
    return rows

def batch_create(data_list, dry_run=True):
    """Batch create product packages with confirmation"""
    print(f"\n{'[DRY RUN] ' if dry_run else ''}准备新建 {len(data_list)} 个商品包：")
    print(f"{'ID':<6} {'名称':<40} {'原价':<12} {'售价':<12}")
    print("-" * 70)
    for i, item in enumerate(data_list):
        print(f"{i+1:<6} {str(item.get('name','?'))[:39]:<40} {item.get('originPrice','?'):<12} {item.get('realPrice','?'):<12}")

    if dry_run:
        print("\n[DRY RUN 模式] 以上是将要新建的数据，请检查后手动调用 batch_create(data, dry_run=False)")
        return

    print("\n确认新建以上商品包？(y/N): ", end="")
    answer = input().strip().lower()
    if answer != "y":
        print("已取消")
        return

    results = {"success": [], "failed": []}
    for i, item in enumerate(data_list):
        payload = {**item, "content": item.get("content", []), "multiCurrencyPricingData": item.get("multiCurrencyPricingData", [])}
        resp = api_post("/international/order-service/api/product/new", payload)
        if resp.get("code") == 200:
            new_id = resp.get("data", {}).get("id", "?")
            print(f"  ✅ [{i+1}] {item.get('name')} → ID={new_id}")
            results["success"].append({"name": item.get("name"), "id": new_id})
        else:
            print(f"  ❌ [{i+1}] {item.get('name')} → {resp.get('msg', resp)}")
            results["failed"].append({"name": item.get("name"), "error": resp.get("msg")})

    print(f"\n完成：成功 {len(results['success'])}，失败 {len(results['failed'])}")
    if results["failed"]:
        print("失败列表：")
        for f in results["failed"]:
            print(f"  {f['name']}: {f['error']}")
    return results

# CLI dispatcher
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    cmd = sys.argv[1]

    if cmd == "auth":
        check_auth()
    elif cmd == "list":
        name = sys.argv[2] if len(sys.argv) > 2 else None
        list_packages(name)
    elif cmd == "detail":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py detail <packageId>")
            sys.exit(1)
        get_detail(int(sys.argv[2]))
    elif cmd == "coupon-limit":
        get_coupon_limit(int(sys.argv[2]))
    elif cmd == "inventory":
        get_inventory(int(sys.argv[2]))
    elif cmd == "update-stock":
        if len(sys.argv) < 4:
            print("用法: python3 ops_helper.py update-stock <packageId> add|subtract|infinity [num]")
            sys.exit(1)
        pkg_id = int(sys.argv[2])
        op = sys.argv[3]
        if op == "infinity":
            update_inventory(pkg_id, None, None)
        elif op in ("add", "subtract"):
            if len(sys.argv) < 5:
                print(f"用法: python3 ops_helper.py update-stock <packageId> {op} <num>")
                sys.exit(1)
            update_inventory(pkg_id, op, int(sys.argv[4]))
        else:
            print(f"[ERROR] 操作类型只支持 add / subtract / infinity，收到：{op}")
    elif cmd == "from-excel":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py from-excel <file.xlsx>")
            sys.exit(1)
        rows = parse_excel(sys.argv[2])
        print(json.dumps(rows[:3], ensure_ascii=False, indent=2, default=str))
    elif cmd == "batch-create":
        if len(sys.argv) < 3:
            print("用法: python3 ops_helper.py batch-create <file.json>")
            sys.exit(1)
        with open(sys.argv[2]) as f:
            data = json.load(f)
        batch_create(data, dry_run=True)
    else:
        print(f"未知命令: {cmd}")
        print(__doc__)
