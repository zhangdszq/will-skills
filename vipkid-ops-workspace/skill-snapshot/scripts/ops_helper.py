#!/usr/bin/env python3
"""
ops_helper.py - VIPKID 运营后台商品包 / 家长详情操作辅助工具
"""
import argparse
import json
import os
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

CONFIG_PATH = Path.home() / ".vipkid-ops" / "config.json"
DEFAULT_BASE_URL = "https://sa-manager.lionabc.com"
DEFAULT_CR_CODE = "sa"
DEFAULT_CDP_PORT = 9222
DEFAULT_REFRESH_MODE = "playwright"
DEFAULT_LOGIN_TIMEOUT_SECONDS = 300
DEFAULT_PLAYWRIGHT_PROFILE_DIR = CONFIG_PATH.parent / "playwright-profile"

PARENT_INFO_FIELDS = {
    "city",
    "hasOverseaStudyExp",
    "ethnicGroup",
    "nationality",
    "monthBudget",
    "memo",
}

LEADS_REFERER_TEMPLATE = "{base_url}/omnicenter/user/parent/{parent_id}?&apiAppCode=international-mgt&headless=1"


def load_config(required=False):
    if not CONFIG_PATH.exists():
        if required:
            print(f"[ERROR] 配置文件不存在：{CONFIG_PATH}")
            print("请先创建配置文件，或直接运行 refresh-token 自动写入 token。")
            print('建议格式：{"base_url": "https://sa-manager.lionabc.com", "token": ""}')
            sys.exit(1)
        return {"base_url": DEFAULT_BASE_URL, "cr_code": DEFAULT_CR_CODE}

    with open(CONFIG_PATH, encoding="utf-8") as file_obj:
        config = json.load(file_obj)

    config.setdefault("base_url", DEFAULT_BASE_URL)
    config.setdefault("cr_code", DEFAULT_CR_CODE)
    return config


def save_config(config):
    CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(CONFIG_PATH, "w", encoding="utf-8") as file_obj:
        json.dump(config, file_obj, indent=2, ensure_ascii=False)
        file_obj.write("\n")


def require_auth_config():
    config = load_config(required=True)
    token = str(config.get("token", "")).strip()
    if token:
        return config

    print(f"[ERROR] 配置文件缺少 token：{CONFIG_PATH}")
    print("可先运行以下命令自动获取：")
    print("  python3 ~/.claude/skills/vipkid-ops/scripts/ops_helper.py refresh-token")
    print("首次运行会打开一个独立 Chrome 窗口，你登录后脚本会自动读取 intlAuthToken。")
    sys.exit(1)


def build_headers(content_type="application/json", header_profile="default", referer=None):
    config = require_auth_config()
    token = config["token"].strip()
    base_url = config["base_url"]
    cr_code = config.get("cr_code", DEFAULT_CR_CODE)

    headers = {
        "intlAuthToken": token,
        "Cookie": f"intlAuthToken={token}",
        "vk-cr-code": cr_code,
        "vk-language-code": "zh-cn",
        "app-code": "international-mgt",
        "Referer": referer or f"{base_url}/hobbit/product/productpackagemanage/",
        "Accept": "application/json",
    }
    if header_profile == "leads":
        headers.update(
            {
                "Authorization": token,
                "Biz-Line": cr_code,
                "App-Code": "international-mgt",
                "Origin": base_url,
                "Accept": "application/json, text/plain, */*",
            }
        )
    if content_type:
        headers["Content-Type"] = content_type
    return config, headers


def api_request(method, path, params=None, data=None, content_type="application/json", header_profile="default", referer=None):
    config, headers = build_headers(content_type=content_type, header_profile=header_profile, referer=referer)
    url = f"{config['base_url']}/rest{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params, doseq=True)

    request_data = None
    if data is not None:
        if content_type == "application/json":
            request_data = json.dumps(data).encode()
        elif content_type == "application/x-www-form-urlencoded":
            request_data = urllib.parse.urlencode(data, doseq=True).encode()
        else:
            request_data = data

    req = urllib.request.Request(url, data=request_data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as error:
        try:
            body = json.loads(error.read().decode())
        except Exception:
            body = {"http_error": error.code, "reason": str(error.reason)}
        return body
    except urllib.error.URLError as error:
        return {"network_error": str(error.reason)}


def api_get(path, params=None, header_profile="default", referer=None):
    return api_request("GET", path, params=params, content_type=None, header_profile=header_profile, referer=referer)


def api_post(path, data, header_profile="default", referer=None):
    return api_request("POST", path, data=data, content_type="application/json", header_profile=header_profile, referer=referer)


def api_post_form(path, data, params=None, header_profile="default", referer=None):
    return api_request("POST", path, params=params, data=data, content_type="application/x-www-form-urlencoded", header_profile=header_profile, referer=referer)


def print_json(data):
    print(json.dumps(data, ensure_ascii=False, indent=2))


def print_api_error(resp):
    print(f"[ERROR] {json.dumps(resp, ensure_ascii=False)}")


def get_response_data(resp):
    return resp.get("data")


def is_success(resp):
    code = resp.get("code")
    error_code = resp.get("errorCode")
    return code in (0, 200) and error_code in (None, 0, "0")


def parent_page_referer(parent_id):
    config = require_auth_config()
    return LEADS_REFERER_TEMPLATE.format(base_url=config["base_url"], parent_id=parent_id)


def parse_refresh_args(args):
    parser = argparse.ArgumentParser(
        prog="python3 ops_helper.py refresh-token",
        description="Refresh intlAuthToken via Playwright or CDP.",
    )
    parser.add_argument(
        "--mode",
        choices=("playwright", "cdp"),
        default=os.environ.get("VIPKID_TOKEN_REFRESH_MODE", DEFAULT_REFRESH_MODE),
        help="Refresh strategy. Defaults to Playwright.",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=int(os.environ.get("VIPKID_CHROME_CDP_PORT", DEFAULT_CDP_PORT)),
        help="CDP port for --mode cdp.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=int(os.environ.get("VIPKID_TOKEN_WAIT_TIMEOUT", DEFAULT_LOGIN_TIMEOUT_SECONDS)),
        help="Seconds to wait for Playwright login.",
    )
    parser.add_argument(
        "--profile-dir",
        default=str(DEFAULT_PLAYWRIGHT_PROFILE_DIR),
        help="Dedicated Playwright profile directory.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Playwright in headless mode. Mainly useful for automated verification.",
    )
    return parser.parse_args(args)


def run_refresh_command(command, timeout_seconds):
    try:
        return subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=False,
        )
    except FileNotFoundError as error:
        print(f"[ERROR] 未检测到命令：{error.filename}")
        return None
    except subprocess.TimeoutExpired:
        print("❌ 自动获取 Token 超时。")
        return None


def refresh_token(mode=DEFAULT_REFRESH_MODE, port=DEFAULT_CDP_PORT, timeout_seconds=DEFAULT_LOGIN_TIMEOUT_SECONDS, profile_dir=DEFAULT_PLAYWRIGHT_PROFILE_DIR, headless=False):
    config = load_config(required=False)
    base_url = config.get("base_url", DEFAULT_BASE_URL)
    profile_dir = Path(profile_dir).expanduser()

    if mode == "playwright":
        script_path = Path(__file__).with_name("get_token_via_playwright.py")
        if not script_path.exists():
            print(f"[ERROR] 未找到 Playwright 脚本：{script_path}")
            return False

        command = [
            sys.executable,
            str(script_path),
            "--url",
            base_url,
            "--cookie",
            "intlAuthToken",
            "--timeout",
            str(timeout_seconds),
            "--profile-dir",
            str(profile_dir),
        ]
        if headless:
            command.append("--headless")

        print(f"正在打开独立 Chrome 窗口，使用专用登录目录：{profile_dir}")
        print("请在弹出的浏览器里完成登录，脚本会在检测到 intlAuthToken 后自动关闭窗口。")
        result = run_refresh_command(command, timeout_seconds + 90)
        failure_hints = [
            "  1. 是否已在弹出的独立 Chrome 窗口中完成登录",
            "  2. 登录成功后 Cookie 中是否存在 intlAuthToken",
            f"  3. 如需重置专用登录态，可删除目录：{profile_dir}",
        ]
        success_message = f"✅ 已通过 Playwright 读取 intlAuthToken，并写入 {CONFIG_PATH}"
    else:
        script_path = Path(__file__).with_name("get_token_via_cdp.js")
        if not script_path.exists():
            print(f"[ERROR] 未找到 CDP 脚本：{script_path}")
            return False

        command = [
            "node",
            str(script_path),
            "--port",
            str(port),
            "--url",
            base_url,
            "--cookie",
            "intlAuthToken",
        ]
        result = run_refresh_command(command, 30)
        failure_hints = [
            f"  1. Chrome 是否已使用 --remote-debugging-port={port} 启动",
            "  2. 是否已在 Chrome 中登录 sa-manager.lionabc.com",
            "  3. Cookie 中是否存在 intlAuthToken",
        ]
        success_message = f"✅ 已通过 Chrome CDP 读取 intlAuthToken，并写入 {CONFIG_PATH}"

    if result is None:
        return False

    token = result.stdout.strip()
    error_output = result.stderr.strip()
    if result.returncode != 0 or not token:
        print("❌ 自动获取 Token 失败。")
        if error_output:
            print(error_output)
        print("可检查：")
        for hint in failure_hints:
            print(hint)
        return False

    config["base_url"] = base_url
    config["cr_code"] = config.get("cr_code", DEFAULT_CR_CODE)
    config["token"] = token
    save_config(config)
    print(success_message)
    return True


def check_auth():
    resp = api_get("/auth/api/auth/userAuthInfo")
    if resp.get("code") == 200 and resp.get("data"):
        role_list = resp["data"].get("roleList", [])
        print(f"✅ 认证有效，角色：{role_list}")
        return True

    print("❌ Token 已过期或无效！")
    if resp.get("network_error"):
        print(f"   网络错误：{resp['network_error']}")
    elif resp.get("http_error"):
        print(f"   HTTP 错误：{resp['http_error']}")
    elif resp.get("msg"):
        print(f"   服务端返回：{resp['msg']}")
    print("   可先尝试自动刷新：")
    print("   python3 ~/.claude/skills/vipkid-ops/scripts/ops_helper.py refresh-token")
    print("   该命令会打开一个独立 Chrome 窗口，你登录后脚本会自动读取 intlAuthToken。")
    print("   若仍失败，再手动从 Chrome → F12 → Application → Cookies → sa-manager.lionabc.com → intlAuthToken 复制 Value")
    return False


def list_packages(name=None, page=0, limit=10):
    params = {"queryType": "Product_Package", "start": page * limit, "limit": limit}
    if name:
        params["productName"] = name
    resp = api_get("/international/api/product/list/", params)
    if resp.get("code") != 200:
        print_api_error(resp)
        return []

    total = resp["data"]["total"]
    items = resp["data"]["data"]
    print(f"共 {total} 个商品包，本页 {len(items)} 个：")
    print(f"{'ID':<8} {'名称':<35} {'状态':<6} {'售价':<15} {'创建人':<12}")
    print("-" * 80)
    for item in items:
        status = "上架" if item.get("status") == 1 else "下架"
        price = item.get("realPrice", "?")
        if isinstance(price, str) and len(price) > 20:
            price = price[:17] + "..."
        print(f"{item['id']:<8} {item['name'][:34]:<35} {status:<6} {str(price):<15} {item.get('createName', '?'):<12}")
    return items


def get_detail(package_id):
    resp = api_get("/international/api/product/detail", {"productId": package_id, "type": "Package"})
    if not is_success(resp):
        print_api_error(resp)
        return None
    data = get_response_data(resp)
    print_json(data)
    return data


def get_coupon_limit(package_id):
    resp = api_get("/international/api/product/getCouponLimit/", {"productId": package_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    data = get_response_data(resp)
    print(f"商品包 {package_id} 优惠券限制：")
    print(f"  数量限制：{data.get('couponLimitNum', '无限制')}")
    print(f"  金额比例限制：{data.get('couponLimitRate', '无限制')}%")
    return data


def set_coupon_limit(package_id, coupon_limit_num="", coupon_limit_rate=""):
    payload = {
        "productId": package_id,
        "couponLimitNum": str(coupon_limit_num or ""),
        "couponLimitRate": str(coupon_limit_rate or ""),
    }
    resp = api_post("/international/order-service/api/product/editCouponLimit/", payload)
    if resp.get("code") == 200:
        print(f"✅ 商品包 {package_id} 优惠券限制已更新")
    else:
        print(f"❌ 更新失败：{resp.get('msg', resp)}")
    return resp


def get_inventory(package_id):
    resp = api_get("/international/order-service/api/product/package/stock", {"packageId": package_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    data = get_response_data(resp)
    print(f"商品包 {package_id} 库存信息：")
    stock_limit = data.get("stockLimit")
    print(f"  是否限制：{'限制' if stock_limit else '不限制'}")
    print(f"  库存总量：{data.get('stockNum', '-')}")
    print(f"  待支付：{data.get('toBePaidNum', '-')}")
    print(f"  退款中：{data.get('inRefundNum', '-')}")
    print(f"  已售：{data.get('paymentNum', '-')}")
    return data


def update_inventory(package_id, operate_type, operate_num, product_type_id=3):
    if operate_type is None:
        payload = {"packageId": package_id, "stockLimit": False}
    elif operate_type in ("add", "subtract"):
        payload = {
            "packageId": package_id,
            "stockLimit": True,
            "operateType": operate_type,
            "operateNum": operate_num,
            "productTypeId": product_type_id,
        }
    else:
        print(f"[ERROR] operateType 只接受 'add' / 'subtract' / None，当前值：{operate_type}")
        return None

    resp = api_post("/international/order-service/api/product/package/updateStock", payload)
    if resp.get("code") == 200:
        print(f"✅ 库存更新成功，当前库存：{resp.get('data', {}).get('stock', '-')}")
    else:
        print(f"❌ 库存更新失败：{resp.get('msg', resp)}")
    return resp


def get_package_auth(package_id):
    resp = api_get("/international/api/productPackage/auth/list", {"packageId": package_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    print_json(get_response_data(resp))
    return get_response_data(resp)


def set_package_auth(package_id, role_ids):
    payload = {"packageId": package_id, "roleIds": role_ids}
    resp = api_post("/international/api/productPackage/auth/config", payload)
    if resp.get("code") == 200:
        print(f"✅ 商品包 {package_id} 赠送权限已更新")
    else:
        print(f"❌ 更新失败：{resp.get('msg', resp)}")
    return resp


def list_roles():
    resp = api_get("/auth/api/auth/role/all")
    if not is_success(resp):
        print_api_error(resp)
        return None
    roles = get_response_data(resp) or []
    print(f"{'ROLE_ID':<10} {'角色名':<40}")
    print("-" * 52)
    for role in roles:
        print(f"{str(role.get('id', '')):<10} {str(role.get('name', '')):<40}")
    return roles


def parse_excel(filepath):
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"Excel 列：{headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(headers, row)))

    print(f"共 {len(rows)} 行数据")
    return rows


def batch_create(data_list, dry_run=True):
    print(f"\n{'[DRY RUN] ' if dry_run else ''}准备新建 {len(data_list)} 个商品包：")
    print(f"{'ID':<6} {'名称':<40} {'原价':<12} {'售价':<12}")
    print("-" * 70)
    for index, item in enumerate(data_list):
        print(f"{index + 1:<6} {str(item.get('name', '?'))[:39]:<40} {item.get('originPrice', '?'):<12} {item.get('realPrice', '?'):<12}")

    if dry_run:
        print("\n[DRY RUN 模式] 以上是将要新建的数据，请检查后带 --execute 再次执行。")
        return {"planned": len(data_list)}

    print("\n确认新建以上商品包？(y/N): ", end="")
    answer = input().strip().lower()
    if answer != "y":
        print("已取消")
        return

    results = {"success": [], "failed": []}
    for index, item in enumerate(data_list):
        payload = {
            **item,
            "content": item.get("content", []),
            "multiCurrencyPricingData": item.get("multiCurrencyPricingData", []),
        }
        resp = api_post("/international/order-service/api/product/new", payload)
        if resp.get("code") == 200:
            new_id = resp.get("data", {}).get("id", resp.get("data", "?"))
            print(f"  ✅ [{index + 1}] {item.get('name')} → ID={new_id}")
            results["success"].append({"name": item.get("name"), "id": new_id})
        else:
            print(f"  ❌ [{index + 1}] {item.get('name')} → {resp.get('msg', resp)}")
            results["failed"].append({"name": item.get("name"), "error": resp.get("msg")})

    print(f"\n完成：成功 {len(results['success'])}，失败 {len(results['failed'])}")
    if results["failed"]:
        print("失败列表：")
        for failed in results["failed"]:
            print(f"  {failed['name']}: {failed['error']}")
    return results


def create_product(payload):
    resp = api_post("/international/order-service/api/product/new", payload)
    print_json(resp)
    return resp


def edit_product(payload):
    if "id" not in payload:
        print("[ERROR] edit 请求必须包含 id")
        return None
    resp = api_post("/international/order-service/api/product/edit", payload)
    print_json(resp)
    return resp


def get_parent_detail(parent_id):
    resp = api_get("/international/api/parent/v2/getParentDetailWithChildById", {"id": parent_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    return get_response_data(resp)


def print_parent_summary(parent_id):
    data = get_parent_detail(parent_id)
    if not data:
        return None
    parent = data.get("parent") or {}
    parent_ext = data.get("parentExt") or {}
    merged = {**parent, **parent_ext}
    children = data.get("childs") or []
    print(f"家长 {parent_id} 概览：")
    print(f"  姓名/昵称：{merged.get('firstName') or merged.get('name') or '-'}")
    print(f"  注册手机号：{merged.get('phone') or '-'}")
    print(f"  联系方式：{merged.get('contactPhone') or '-'}")
    print(f"  邮箱：{merged.get('email') or '-'}")
    print(f"  邀请码：{merged.get('referralCode') or '-'}")
    print(f"  推荐人ID：{merged.get('referrerId') or '-'}")
    print(f"  时区：{merged.get('timeZone') or '-'}")
    print(f"  状态：{merged.get('status') or '-'}")
    print(f"  注册渠道：{merged.get('channelName') or '-'}")
    print(f"  注册来源：{merged.get('deviceType') or '-'}")
    print(f"  学生数：{len(children)}")
    if children:
        print("  学生列表：")
        for child in children:
            print(f"    SID={child.get('id')} name={child.get('name')} englishName={child.get('englishName', '')}")
    return {"parent": merged, "childs": children}


def update_parent_info(parent_id, field, value):
    if field not in PARENT_INFO_FIELDS:
        print(f"[ERROR] 不支持字段：{field}")
        print(f"支持字段：{', '.join(sorted(PARENT_INFO_FIELDS))}")
        return None
    resp = api_post_form("/international/uc/api/parent/v2/updateInfo", {"id": parent_id, field: value})
    print_json(resp)
    return resp


def update_parent_contacts(parent_id, contacts):
    payload = {
        "parentId": parent_id,
        "otherContacts": json.dumps([{"data": item} for item in contacts], ensure_ascii=False),
    }
    resp = api_post_form("/international/api/parent/editOtherContacts", payload)
    print_json(resp)
    return resp


def update_parent_timezone(parent_id, timezone_value):
    resp = api_post_form("/international/api/parent/updateTimeZone", {"parentId": parent_id, "timeZone": timezone_value})
    print_json(resp)
    return resp


def update_parent_referrer(parent_id, referrer_code):
    resp = api_post_form("/international/api/parent/updateReferrerInfo", {"parentId": parent_id, "referrerCode": referrer_code})
    print_json(resp)
    return resp


def decrypt_parent(parent_id):
    resp = api_get("/international/api/parent/decrypt", {"id": parent_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    print_json(get_response_data(resp))
    return get_response_data(resp)


def get_user_has_login(parent_id, device_type="APP"):
    resp = api_get("/international/uc/api/rpc/user/hasLogin", {"userId": parent_id, "deviceType": device_type})
    if not is_success(resp):
        print_api_error(resp)
        return None
    print_json(get_response_data(resp))
    return get_response_data(resp)


def get_class_progress(parent_id):
    resp = api_get("/manager/webhub/admin/api/admin/class/progress", {"parentId": parent_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    data = get_response_data(resp)
    print_json(data)
    return data


def get_minimum_consumption_detail(consumption_id):
    resp = api_get(f"/international/api/minimumConsumption/schedule/detail/{consumption_id}", {"id": consumption_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    print_json(get_response_data(resp))
    return get_response_data(resp)


def get_batch_learning_behavior(parent_id):
    resp = api_get("/manager/webhub/admin/api/class/schedule/batchLearningBehavior", {"parentId": parent_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    data = get_response_data(resp)
    print_json(data)
    return data


def get_learning_behavior(student_id):
    resp = api_get("/manager/webhub/admin/api/class/schedule/learningBehavior", {"studentId": student_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    data = get_response_data(resp)
    print_json(data)
    return data


def get_follow_records(parent_id):
    student_resp = get_parent_detail(parent_id) or {}
    students = student_resp.get("childs") or []

    node_follow_map = {}
    for child in students:
        student_id = child.get("id")
        if not student_id:
            continue
        resp = api_get("/international/api/flow/getFlowRemarkList", {"studentId": student_id, "type": 0})
        if is_success(resp):
            node_follow_map[str(student_id)] = get_response_data(resp) or []
        else:
            node_follow_map[str(student_id)] = {"error": resp}

    sale_resp = api_get("/international/api/sale/tracking/list", {"parentId": parent_id})
    leads_referer = parent_page_referer(parent_id)
    service_resp = api_get(
        "/international/leads/api/v2/leads/follow/record/getList",
        {"userId": parent_id},
        header_profile="leads",
        referer=leads_referer,
    )
    sobot_resp = api_post(
        "/international/leads/api/v1/leads/sobot/callRecord/search/?userId=%s" % parent_id,
        {},
        header_profile="leads",
        referer=leads_referer,
    )

    result = {
        "students": students,
        "node_follow_records": node_follow_map,
        "sale_tracking_records": (get_response_data(sale_resp) or {}).get("data", []) if is_success(sale_resp) else sale_resp,
        "service_follow_records": get_response_data(service_resp) or [] if is_success(service_resp) else service_resp,
        "sobot_call_records": get_response_data(sobot_resp) or [] if is_success(sobot_resp) else sobot_resp,
    }
    print_json(result)
    return result


def get_poster_config():
    resp = api_get("/international/api/poster/config/getInfo")
    if not is_success(resp):
        print_api_error(resp)
        return None
    print_json(get_response_data(resp))
    return get_response_data(resp)


def get_busiclick_total(parent_id, action="posterCopy"):
    resp = api_get("/international/api/busiclick/getTotal", {"action": action, "parentId": parent_id})
    if not is_success(resp):
        print_api_error(resp)
        return None
    print_json(get_response_data(resp))
    return get_response_data(resp)


def build_parser():
    parser = argparse.ArgumentParser(description="VIPKID 运营后台辅助工具")
    subparsers = parser.add_subparsers(dest="command")

    subparsers.add_parser("auth", help="验证 token 是否有效")

    refresh_parser = subparsers.add_parser("refresh-token", help="自动刷新 token")
    refresh_parser.add_argument("--mode", choices=("playwright", "cdp"), default=DEFAULT_REFRESH_MODE)
    refresh_parser.add_argument("--port", type=int, default=DEFAULT_CDP_PORT)
    refresh_parser.add_argument("--timeout", type=int, default=DEFAULT_LOGIN_TIMEOUT_SECONDS)
    refresh_parser.add_argument("--profile-dir", default=str(DEFAULT_PLAYWRIGHT_PROFILE_DIR))
    refresh_parser.add_argument("--headless", action="store_true")

    list_parser = subparsers.add_parser("list", help="列出商品包")
    list_parser.add_argument("name", nargs="?")
    list_parser.add_argument("--page", type=int, default=0)
    list_parser.add_argument("--limit", type=int, default=10)

    detail_parser = subparsers.add_parser("detail", help="商品包详情")
    detail_parser.add_argument("package_id", type=int)

    coupon_parser = subparsers.add_parser("coupon-limit", help="查看商品包优惠券限制")
    coupon_parser.add_argument("package_id", type=int)

    set_coupon_parser = subparsers.add_parser("set-coupon-limit", help="修改商品包优惠券限制")
    set_coupon_parser.add_argument("package_id", type=int)
    set_coupon_parser.add_argument("--num", default="")
    set_coupon_parser.add_argument("--rate", default="")

    inventory_parser = subparsers.add_parser("inventory", help="查看库存")
    inventory_parser.add_argument("package_id", type=int)

    stock_parser = subparsers.add_parser("update-stock", help="更新库存")
    stock_parser.add_argument("package_id", type=int)
    stock_parser.add_argument("operation", choices=("add", "subtract", "infinity"))
    stock_parser.add_argument("num", nargs="?", type=int)
    stock_parser.add_argument("--product-type-id", type=int, default=3)

    auth_list_parser = subparsers.add_parser("package-auth", help="查看商品包赠送权限")
    auth_list_parser.add_argument("package_id", type=int)

    auth_set_parser = subparsers.add_parser("set-package-auth", help="配置商品包赠送权限")
    auth_set_parser.add_argument("package_id", type=int)
    auth_set_parser.add_argument("role_ids", nargs="+", type=int)

    subparsers.add_parser("roles", help="列出全部角色")

    from_excel_parser = subparsers.add_parser("from-excel", help="解析 Excel")
    from_excel_parser.add_argument("file")

    batch_create_parser = subparsers.add_parser("batch-create", help="批量新建商品包")
    batch_create_parser.add_argument("file")
    batch_create_parser.add_argument("--execute", action="store_true")

    create_parser = subparsers.add_parser("create-product", help="用 JSON 文件新建商品包")
    create_parser.add_argument("file")

    edit_parser = subparsers.add_parser("edit-product", help="用 JSON 文件修改商品包")
    edit_parser.add_argument("file")

    parent_parser = subparsers.add_parser("parent-detail", help="查看家长详情页聚合数据")
    parent_parser.add_argument("parent_id", type=int)
    parent_parser.add_argument("--json", action="store_true")

    parent_update_parser = subparsers.add_parser("update-parent-info", help="更新家长资料字段")
    parent_update_parser.add_argument("parent_id", type=int)
    parent_update_parser.add_argument("field")
    parent_update_parser.add_argument("value")

    parent_contacts_parser = subparsers.add_parser("update-parent-contacts", help="更新备注联系方式")
    parent_contacts_parser.add_argument("parent_id", type=int)
    parent_contacts_parser.add_argument("contacts", nargs="+")

    parent_timezone_parser = subparsers.add_parser("update-parent-timezone", help="更新家长时区")
    parent_timezone_parser.add_argument("parent_id", type=int)
    parent_timezone_parser.add_argument("timezone")

    parent_referrer_parser = subparsers.add_parser("update-parent-referrer", help="补填推荐码")
    parent_referrer_parser.add_argument("parent_id", type=int)
    parent_referrer_parser.add_argument("referrer_code")

    parent_decrypt_parser = subparsers.add_parser("decrypt-parent", help="解密查看家长联系方式")
    parent_decrypt_parser.add_argument("parent_id", type=int)

    has_login_parser = subparsers.add_parser("has-login", help="查询家长是否登录过某设备")
    has_login_parser.add_argument("parent_id", type=int)
    has_login_parser.add_argument("--device-type", default="APP")

    class_progress_parser = subparsers.add_parser("class-progress", help="查看家长课耗进度")
    class_progress_parser.add_argument("parent_id", type=int)

    min_consume_parser = subparsers.add_parser("minimum-consumption-detail", help="查看固定课消明细")
    min_consume_parser.add_argument("consumption_id", type=int)

    batch_behavior_parser = subparsers.add_parser("batch-learning-behavior", help="查看家长下全部学生学习行为")
    batch_behavior_parser.add_argument("parent_id", type=int)

    behavior_parser = subparsers.add_parser("learning-behavior", help="查看单个学生学习行为")
    behavior_parser.add_argument("student_id", type=int)

    follow_parser = subparsers.add_parser("follow-records", help="查看家长详情页跟进记录聚合")
    follow_parser.add_argument("parent_id", type=int)

    subparsers.add_parser("poster-config", help="查看邀请海报配置")

    poster_count_parser = subparsers.add_parser("poster-copy-count", help="查看海报复制次数")
    poster_count_parser.add_argument("parent_id", type=int)
    poster_count_parser.add_argument("--action", default="posterCopy")

    return parser


def main():
    parser = build_parser()
    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return 0

    if args.command == "auth":
        check_auth()
    elif args.command == "refresh-token":
        refresh_token(
            mode=args.mode,
            port=args.port,
            timeout_seconds=args.timeout,
            profile_dir=args.profile_dir,
            headless=args.headless,
        )
    elif args.command == "list":
        list_packages(args.name, page=args.page, limit=args.limit)
    elif args.command == "detail":
        get_detail(args.package_id)
    elif args.command == "coupon-limit":
        get_coupon_limit(args.package_id)
    elif args.command == "set-coupon-limit":
        set_coupon_limit(args.package_id, args.num, args.rate)
    elif args.command == "inventory":
        get_inventory(args.package_id)
    elif args.command == "update-stock":
        if args.operation == "infinity":
            update_inventory(args.package_id, None, None, product_type_id=args.product_type_id)
        else:
            if args.num is None:
                print(f"[ERROR] 操作 {args.operation} 需要 num")
                return 1
            update_inventory(args.package_id, args.operation, args.num, product_type_id=args.product_type_id)
    elif args.command == "package-auth":
        get_package_auth(args.package_id)
    elif args.command == "set-package-auth":
        set_package_auth(args.package_id, args.role_ids)
    elif args.command == "roles":
        list_roles()
    elif args.command == "from-excel":
        rows = parse_excel(args.file)
        print_json(rows[:3])
    elif args.command == "batch-create":
        with open(args.file, encoding="utf-8") as file_obj:
            data = json.load(file_obj)
        batch_create(data, dry_run=not args.execute)
    elif args.command == "create-product":
        with open(args.file, encoding="utf-8") as file_obj:
            create_product(json.load(file_obj))
    elif args.command == "edit-product":
        with open(args.file, encoding="utf-8") as file_obj:
            edit_product(json.load(file_obj))
    elif args.command == "parent-detail":
        if args.json:
            print_json(get_parent_detail(args.parent_id))
        else:
            print_parent_summary(args.parent_id)
    elif args.command == "update-parent-info":
        update_parent_info(args.parent_id, args.field, args.value)
    elif args.command == "update-parent-contacts":
        update_parent_contacts(args.parent_id, args.contacts)
    elif args.command == "update-parent-timezone":
        update_parent_timezone(args.parent_id, args.timezone)
    elif args.command == "update-parent-referrer":
        update_parent_referrer(args.parent_id, args.referrer_code)
    elif args.command == "decrypt-parent":
        decrypt_parent(args.parent_id)
    elif args.command == "has-login":
        get_user_has_login(args.parent_id, device_type=args.device_type)
    elif args.command == "class-progress":
        get_class_progress(args.parent_id)
    elif args.command == "minimum-consumption-detail":
        get_minimum_consumption_detail(args.consumption_id)
    elif args.command == "batch-learning-behavior":
        get_batch_learning_behavior(args.parent_id)
    elif args.command == "learning-behavior":
        get_learning_behavior(args.student_id)
    elif args.command == "follow-records":
        get_follow_records(args.parent_id)
    elif args.command == "poster-config":
        get_poster_config()
    elif args.command == "poster-copy-count":
        get_busiclick_total(args.parent_id, action=args.action)
    else:
        parser.print_help()
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
