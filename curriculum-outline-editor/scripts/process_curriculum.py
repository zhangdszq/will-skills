#!/usr/bin/env python3
"""
Curriculum Excel processor for curriculum-outline-editor skill.

Usage:
  Read mode:   python3 process_curriculum.py read <filepath>
  Write mode:  python3 process_curriculum.py write <original> <changes_json> <output>

Changes JSON format:
[
  {
    "sheet": "Sheet1",
    "row": 3,
    "col": 2,
    "new_value": "Stand up!"
  },
  ...
]
"""

import sys
import json
import os

def read_excel(filepath):
    """Read all sheets and output structured JSON."""
    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {"file": os.path.basename(filepath), "sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            row_data = []
            for col_idx, cell in enumerate(row, start=1):
                # MergedCell objects don't have column_letter; use openpyxl utils
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_idx)
                row_data.append({
                    "row": row_idx,
                    "col": col_idx,
                    "col_letter": col_letter,
                    "value": cell.value
                })
            # Skip entirely empty rows
            if any(c["value"] is not None for c in row_data):
                rows.append(row_data)
        result["sheets"][sheet_name] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "rows": rows
        }

    print(json.dumps(result, ensure_ascii=False, indent=2))


def write_excel(original_path, changes_json_path, output_path):
    """Apply changes from JSON and write a new Excel file."""
    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
        sys.exit(1)

    with open(changes_json_path, "r", encoding="utf-8") as f:
        changes = json.load(f)

    wb = openpyxl.load_workbook(original_path)
    applied = []
    errors = []

    for change in changes:
        sheet_name = change.get("sheet")
        row = change.get("row")
        col = change.get("col")
        new_value = change.get("new_value")

        if sheet_name not in wb.sheetnames:
            errors.append(f"Sheet '{sheet_name}' not found")
            continue

        ws = wb[sheet_name]
        old_value = ws.cell(row=row, column=col).value
        ws.cell(row=row, column=col).value = new_value
        applied.append({
            "sheet": sheet_name,
            "row": row,
            "col": col,
            "old": old_value,
            "new": new_value
        })

    wb.save(output_path)
    print(json.dumps({
        "success": True,
        "output": output_path,
        "applied_count": len(applied),
        "error_count": len(errors),
        "applied": applied,
        "errors": errors
    }, ensure_ascii=False, indent=2))


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    mode = sys.argv[1]

    if mode == "read":
        read_excel(sys.argv[2])
    elif mode == "write":
        if len(sys.argv) < 5:
            print("write mode requires: <original> <changes_json> <output>")
            sys.exit(1)
        write_excel(sys.argv[2], sys.argv[3], sys.argv[4])
    else:
        print(f"Unknown mode: {mode}. Use 'read' or 'write'.")
        sys.exit(1)


if __name__ == "__main__":
    main()
